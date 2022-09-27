# ! python3.8.10

"""
    readCensusExcel.py - Tabulates population and number of census tracks for each country
"""
import openpyxl
import pprint

print('Opening workbook...')

wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb['Population by Census Tract']
countryData = {}

print("Reading rows...")

for row in range(2, sheet.max_row + 1):  # not sure about sheet.max_row
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value
    countryData.setdefault(state, {})
    countryData[state].setdefault(county, {'tracks': 0, 'pop': 0})
    countryData[state][county]['tracks'] += 1
    countryData[state][county]['pop'] += int(pop)

print('Writing results ...')
resultFile = open('census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(countryData))
resultFile.close()

print('***Done***')
